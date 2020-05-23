import openpyxl
import os
from .models import words

class load_excel(object):
    def __init__(self, filedir = ""):
        try:
            self.wb = openpyxl.load_workbook(filedir)
        except Exception as e:
            print(e.args)
    
    def get_content(self, Product):
        try:
            sheetnames = self.wb.sheetnames
            for name in sheetnames:
                sh = self.wb[name]
                for row in range(1, sh.max_row + 1):
                    if sh.cell(row, 2).value == None:
                        continue
                    content = []
                    for col in range(2, sh.max_column + 1):
                        if sh.cell(row, col).value == None:
                            continue
                        content.append(sh.cell(row, col).value)
                    if len(content) != 4:
                        continue
                    words_set = words.objects.filter(word = content[0])
                    if words_set.exists():
                        word = words_set.first()
                        if word.product.filter(id = Product.id).exists():
                            continue
                        word.product.add(Product)
                        word.save()
                    else:
                        word = words()
                        word.word = content[0]
                        word.symthm = content[1]
                        word.chinese = content[2]
                        word.analyzation = content[3]
                        word.product.add(Product)
                        word.save()
        except Exception as e:
            print(e.args)

if __name__ == "__main__":
    Base_dir = os.path.dirname(__file__)
    filedir = os.path.join(Base_dir, 'static\words')
    excel = load_excel(filedir + r"\words.xlsx")
    excel.get_content()
